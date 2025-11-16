#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
指定された xlsx のテスト結果を読み取り、FAIL/ERROR の行を出力する補助スクリプト
Usage:
    python scripts\read_test_xlsx.py path/to/test_results.xlsx
"""
import sys
import os

path = sys.argv[1] if len(sys.argv) > 1 else None
if not path or not os.path.exists(path):
    print('File not found:', path)
    sys.exit(2)

try:
    from openpyxl import load_workbook
except Exception as e:
    print('openpyxl not available:', e)
    sys.exit(3)

wb = load_workbook(path, read_only=True)
if 'Test Results' in wb.sheetnames:
    ws = wb['Test Results']
else:
    ws = wb.active

rows = list(ws.iter_rows(values_only=True))
# header assumed at row 0
header = rows[0] if rows else []
print('Header:', header)

failures = []
for r in rows[1:]:
    cls, test, status, details = (r + ('', '', '', ''))[:4]
    if status in ('FAIL', 'ERROR'):
        failures.append({'class': cls, 'test': test, 'status': status, 'details': details})

if not failures:
    print('No FAIL/ERROR entries found.')
else:
    print(f'Found {len(failures)} failure(s):')
    for i, f in enumerate(failures, 1):
        print('---')
        print(f"{i}. {f['class']}.{f['test']} -> {f['status']}")
        print(f['details'][:2000])

# summary sheet
if 'Summary' in wb.sheetnames:
    ws2 = wb['Summary']
    print('\nSummary:')
    for row in ws2.iter_rows(values_only=True):
        print(row)
else:
    print('\nNo Summary sheet')
